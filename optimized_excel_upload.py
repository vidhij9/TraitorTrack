"""
Optimized Excel Upload Handler for TraceTrack
Handles 80,000+ bags with ultra-fast batch processing
"""
import os
import time
import logging
from io import BytesIO
from collections import defaultdict, deque
from typing import Dict, List, Set, Tuple, Optional
import openpyxl
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values, execute_batch
from flask import current_app
from sqlalchemy import text
from app_clean import db
from models import Bag, Link, Scan, BagType
from datetime import datetime

logger = logging.getLogger(__name__)

class OptimizedExcelUploader:
    """Ultra-optimized Excel upload handler for massive datasets"""
    
    def __init__(self):
        self.batch_size = 2000  # Reduced batch size for large files (60k-80k+ rows) to prevent timeouts
        self.chunk_size = 10000  # Excel reading chunk size
        self.max_memory_mb = 500  # Maximum memory usage in MB
        self.database_url = os.environ.get('DATABASE_URL')
        self.stats = {
            'total_rows': 0,
            'successful_links': 0,
            'parent_bags_created': 0,
            'child_bags_created': 0,
            'existing_links': 0,
            'duplicate_children': 0,
            'invalid_format': 0,
            'errors': []
        }
        
    def process_excel_file(self, file_content: bytes, user_id: int, dispatch_area: str) -> Dict:
        """
        Process Excel file with 80k+ bags efficiently
        Accepts any format of parent and child bags
        """
        start_time = time.time()
        logger.info("Starting optimized Excel processing")
        
        try:
            # Reset statistics
            self.stats = {
                'total_rows': 0,
                'successful_links': 0,
                'parent_bags_created': 0,
                'child_bags_created': 0,
                'existing_links': 0,
                'duplicate_children': 0,
                'invalid_format': 0,
                'errors': []
            }
            
            # Step 1: Parse Excel and collect unique pairs
            parent_child_pairs = self._parse_excel_optimized(file_content)
            
            if not parent_child_pairs:
                logger.warning("No valid data found in Excel file")
                return self.stats
            
            # Step 2: Process bags in optimized batches
            self._process_bags_batch(parent_child_pairs, user_id, dispatch_area)
            
            elapsed = time.time() - start_time
            logger.info(f"Excel processing completed in {elapsed:.2f} seconds")
            logger.info(f"Stats: {self.stats}")
            
            return self.stats
            
        except Exception as e:
            logger.error(f"Excel processing error: {str(e)}")
            self.stats['errors'].append(f"Processing error: {str(e)}")
            return self.stats
    
    def _parse_excel_optimized(self, file_content: bytes) -> List[Tuple[str, str]]:
        """
        Parse Excel file with streaming to handle large files efficiently
        Returns list of unique (parent_qr, child_qr) pairs
        """
        logger.info("Parsing Excel file")
        
        try:
            # Use read_only mode for better memory efficiency
            workbook = load_workbook(
                BytesIO(file_content), 
                read_only=True, 
                data_only=True,
                keep_links=False
            )
            sheet = workbook.active
            
            # Track unique combinations to avoid duplicates
            seen_pairs = set()
            unique_pairs = []
            
            # Process in chunks for memory efficiency
            row_buffer = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 3:
                    continue
                
                # Extract parent and child QR codes (columns B and C)
                # Accept any format - no validation on format
                parent_qr = str(row[1]).strip().upper() if row[1] else None
                child_qr = str(row[2]).strip().upper() if row[2] else None
                
                if not parent_qr or not child_qr:
                    self.stats['invalid_format'] += 1
                    continue
                
                # Remove any whitespace or special characters
                parent_qr = ''.join(parent_qr.split())
                child_qr = ''.join(child_qr.split())
                
                # Check for duplicates
                pair = (parent_qr, child_qr)
                if pair in seen_pairs:
                    self.stats['duplicate_children'] += 1
                    continue
                
                seen_pairs.add(pair)
                row_buffer.append(pair)
                self.stats['total_rows'] += 1
                
                # Process buffer when it reaches chunk size
                if len(row_buffer) >= self.chunk_size:
                    unique_pairs.extend(row_buffer)
                    row_buffer = []
                    logger.info(f"Processed {self.stats['total_rows']} rows")
            
            # Add remaining buffer
            if row_buffer:
                unique_pairs.extend(row_buffer)
            
            workbook.close()
            logger.info(f"Parsed {len(unique_pairs)} unique parent-child pairs")
            
            return unique_pairs
            
        except Exception as e:
            logger.error(f"Excel parsing error: {str(e)}")
            self.stats['errors'].append(f"File parsing error: {str(e)}")
            return []
    
    def _process_bags_batch(self, parent_child_pairs: List[Tuple[str, str]], 
                           user_id: int, dispatch_area: str):
        """
        Process bags in optimized batches with bulk operations
        """
        logger.info(f"Processing {len(parent_child_pairs)} parent-child pairs")
        
        # Extract unique parent and child QR codes
        parent_qrs = set()
        child_qrs = set()
        
        for parent_qr, child_qr in parent_child_pairs:
            parent_qrs.add(parent_qr)
            child_qrs.add(child_qr)
        
        logger.info(f"Unique parents: {len(parent_qrs)}, Unique children: {len(child_qrs)}")
        
        # Get direct database connection for bulk operations
        conn = psycopg2.connect(self.database_url)
        cur = conn.cursor()
        
        try:
            # Step 1: Load existing bags in batches
            existing_parents = self._load_existing_bags_batch(cur, parent_qrs)
            existing_children = self._load_existing_bags_batch(cur, child_qrs)
            
            # Step 2: Create missing bags
            new_parents = parent_qrs - set(existing_parents.keys())
            new_children = child_qrs - set(existing_children.keys())
            
            if new_parents:
                self._create_bags_bulk(cur, new_parents, 'parent', user_id, dispatch_area)
                self.stats['parent_bags_created'] = len(new_parents)
            
            if new_children:
                self._create_bags_bulk(cur, new_children, 'child', user_id, dispatch_area)
                self.stats['child_bags_created'] = len(new_children)
            
            # Step 3: Reload all bags with IDs
            all_parents = self._load_existing_bags_batch(cur, parent_qrs)
            all_children = self._load_existing_bags_batch(cur, child_qrs)
            
            # Step 4: Load existing links
            existing_links = self._load_existing_links_batch(cur, all_parents, all_children)
            
            # Step 5: Create new links and scans
            self._create_links_and_scans_bulk(
                cur, parent_child_pairs, all_parents, all_children, 
                existing_links, user_id
            )
            
            # Step 6: Update parent bag counts and weights
            self._update_parent_counts_bulk(cur, all_parents.values())
            
            # Commit all changes
            conn.commit()
            logger.info("All database operations committed successfully")
            
        except Exception as e:
            conn.rollback()
            logger.error(f"Database operation error: {str(e)}")
            self.stats['errors'].append(f"Database error: {str(e)}")
            raise
        finally:
            cur.close()
            conn.close()
    
    def _load_existing_bags_batch(self, cur, qr_ids: Set[str]) -> Dict[str, int]:
        """
        Load existing bags in batches, returns {qr_id: bag_id}
        """
        if not qr_ids:
            return {}
        
        result = {}
        qr_list = list(qr_ids)
        
        # Process in batches to avoid query size limits
        for i in range(0, len(qr_list), self.batch_size):
            batch = qr_list[i:i + self.batch_size]
            
            cur.execute("""
                SELECT id, qr_id FROM bag 
                WHERE qr_id = ANY(%s)
            """, (batch,))
            
            for bag_id, qr_id in cur.fetchall():
                result[qr_id] = bag_id
        
        return result
    
    def _create_bags_bulk(self, cur, qr_ids: Set[str], bag_type: str, 
                         user_id: int, dispatch_area: str):
        """
        Create bags in bulk using COPY for maximum performance
        """
        if not qr_ids:
            return
        
        logger.info(f"Creating {len(qr_ids)} {bag_type} bags")
        
        # Prepare data for bulk insert
        now = datetime.utcnow()
        bags_data = []
        
        for qr_id in qr_ids:
            bags_data.append((
                qr_id,
                bag_type,
                'pending',
                user_id,
                dispatch_area,
                0 if bag_type == 'parent' else None,  # child_count
                0.0 if bag_type == 'parent' else 1.0,  # weight_kg
                now,
                now
            ))
        
        # Use execute_values for efficient bulk insert
        execute_values(
            cur,
            """
            INSERT INTO bag (qr_id, type, status, user_id, dispatch_area, 
                           child_count, weight_kg, created_at, updated_at)
            VALUES %s
            ON CONFLICT (qr_id) DO NOTHING
            """,
            bags_data,
            template="(%s, %s, %s, %s, %s, %s, %s, %s, %s)"
        )
    
    def _load_existing_links_batch(self, cur, parent_bags: Dict[str, int], 
                                  child_bags: Dict[str, int]) -> Set[Tuple[int, int]]:
        """
        Load existing links between bags
        """
        if not parent_bags or not child_bags:
            return set()
        
        parent_ids = list(parent_bags.values())
        child_ids = list(child_bags.values())
        
        cur.execute("""
            SELECT parent_bag_id, child_bag_id 
            FROM link 
            WHERE parent_bag_id = ANY(%s) 
            AND child_bag_id = ANY(%s)
        """, (parent_ids, child_ids))
        
        return {(row[0], row[1]) for row in cur.fetchall()}
    
    def _create_links_and_scans_bulk(self, cur, parent_child_pairs: List[Tuple[str, str]],
                                    parent_bags: Dict[str, int], child_bags: Dict[str, int],
                                    existing_links: Set[Tuple[int, int]], user_id: int):
        """
        Create links and scans in bulk
        """
        links_to_create = []
        scans_to_create = []
        now = datetime.utcnow()
        
        for parent_qr, child_qr in parent_child_pairs:
            parent_id = parent_bags.get(parent_qr)
            child_id = child_bags.get(child_qr)
            
            if not parent_id or not child_id:
                continue
            
            # Check if link exists
            if (parent_id, child_id) in existing_links:
                self.stats['existing_links'] += 1
            else:
                # Add new link
                links_to_create.append((parent_id, child_id, now))
                self.stats['successful_links'] += 1
            
            # Always create scan record for audit
            scans_to_create.append((parent_id, child_id, user_id, now))
        
        # Bulk insert links in smaller batches to avoid timeouts
        if links_to_create:
            # Process in batches of 5000 for very large uploads
            LINK_BATCH_SIZE = 5000
            for i in range(0, len(links_to_create), LINK_BATCH_SIZE):
                batch = links_to_create[i:i + LINK_BATCH_SIZE]
                execute_values(
                    cur,
                    """
                    INSERT INTO link (parent_bag_id, child_bag_id, created_at)
                    VALUES %s
                    ON CONFLICT (parent_bag_id, child_bag_id) DO NOTHING
                    """,
                    batch,
                    page_size=500  # Smaller page size for stability
                )
                # Commit periodically for very large batches
                if i > 0 and i % 20000 == 0:
                    cur.connection.commit()
                    logger.info(f"Committed {i} links...")
        
        # Bulk insert scans in batches
        if scans_to_create:
            SCAN_BATCH_SIZE = 5000
            for i in range(0, len(scans_to_create), SCAN_BATCH_SIZE):
                batch = scans_to_create[i:i + SCAN_BATCH_SIZE]
                execute_values(
                    cur,
                    """
                    INSERT INTO scan (parent_bag_id, child_bag_id, user_id, timestamp)
                    VALUES %s
                    """,
                    batch,
                    page_size=500
                )
                # Commit periodically
                if i > 0 and i % 20000 == 0:
                    cur.connection.commit()
                    logger.info(f"Committed {i} scans...")
    
    def _update_parent_counts_bulk(self, cur, parent_ids: List[int]):
        """
        Update parent bag counts and weights in bulk
        """
        if not parent_ids:
            return
        
        logger.info(f"Updating counts for {len(parent_ids)} parent bags")
        
        # Update all parent bags in one query
        cur.execute("""
            UPDATE bag p
            SET 
                child_count = COALESCE((
                    SELECT COUNT(*) 
                    FROM link l 
                    WHERE l.parent_bag_id = p.id
                ), 0),
                weight_kg = COALESCE((
                    SELECT COUNT(*) 
                    FROM link l 
                    WHERE l.parent_bag_id = p.id
                ), 0)::float,
                status = CASE 
                    WHEN (SELECT COUNT(*) FROM link l WHERE l.parent_bag_id = p.id) >= 30 
                    THEN 'completed'
                    ELSE 'pending'
                END,
                updated_at = NOW()
            WHERE p.id = ANY(%s) AND p.type = 'parent'
        """, (list(parent_ids),))


def create_test_excel(filename: str, num_parents: int = 2667, children_per_parent: int = 30):
    """
    Create a test Excel file with specified number of bags
    Default: 2667 parents * 30 children = 80,010 bags
    """
    from xlsxwriter import Workbook
    
    workbook = Workbook(filename)
    worksheet = workbook.add_worksheet()
    
    # Write headers
    worksheet.write(0, 0, 'Serial')
    worksheet.write(0, 1, 'Parent Bag')
    worksheet.write(0, 2, 'Child Bag')
    
    row = 1
    serial = 1
    
    for parent_num in range(num_parents):
        # Generate parent QR with various formats
        if parent_num % 3 == 0:
            parent_qr = f"SB{parent_num:05d}"
        elif parent_num % 3 == 1:
            parent_qr = f"PB{parent_num:06d}"
        else:
            parent_qr = f"BAG{parent_num:04d}"
        
        for child_num in range(children_per_parent):
            # Generate child QR with various formats
            child_id = parent_num * children_per_parent + child_num
            
            if child_num % 4 == 0:
                child_qr = f"{child_id:06d}"
            elif child_num % 4 == 1:
                child_qr = f"CB{child_id:05d}"
            elif child_num % 4 == 2:
                child_qr = f"C-{child_id}"
            else:
                child_qr = f"CHILD{child_id}"
            
            worksheet.write(row, 0, serial)
            worksheet.write(row, 1, parent_qr)
            worksheet.write(row, 2, child_qr)
            
            row += 1
            serial += 1
            
            # Show progress every 10000 rows
            if row % 10000 == 0:
                print(f"Written {row} rows...")
    
    workbook.close()
    print(f"Test Excel file created: {filename}")
    print(f"Total rows: {row - 1}")
    print(f"Parents: {num_parents}, Children per parent: {children_per_parent}")
    
    return filename


# Export the main class
excel_uploader = OptimizedExcelUploader()