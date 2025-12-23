"""
Bulk Import Utilities for TraitorTrack
Provides CSV and Excel import functionality with comprehensive validation

FEATURES:
- Import bags (parent and child) with validation
- Import bills with parent bag linkages
- Duplicate detection and handling
- Error reporting with line numbers
- Batch processing for performance
- Admin-only access enforced in routes
- STREAMING Excel processing for large files (100k+ rows)
- Per-row status tracking for detailed result files
- Memory-efficient chunked processing

SAFETY:
- Input validation for all fields
- Duplicate prevention
- Transaction-based imports (rollback on error)
- Memory-efficient batch processing with streaming
- SAVEPOINT-based batch rollback for resilience
"""
import csv
import io
import logging
import gc
import os
import uuid
import tempfile
from datetime import datetime
from typing import List, Dict, Tuple, Optional, Generator, Any, Callable
from flask import flash
from sqlalchemy import text
from werkzeug.datastructures import FileStorage
from app import db
from models import Bag, Bill, Link, BillBag, User

logger = logging.getLogger(__name__)

# Configuration for large file handling
CHUNK_SIZE = 2000  # Rows per database commit batch
MAX_ERRORS_PER_FILE = 1000  # Limit error collection to prevent memory issues
STREAMING_THRESHOLD = 10000  # Use streaming for files with more rows

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel imports disabled")


class ImportValidator:
    """Validates import data before database insertion"""
    
    @staticmethod
    def validate_bag_data(row: Dict, row_num: int) -> Tuple[bool, Optional[str]]:
        """
        Validate a single bag row.
        
        Args:
            row: Dictionary with bag data
            row_num: Row number for error reporting
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        # Required fields
        required_fields = ['qr_id', 'type']
        for field in required_fields:
            if not row.get(field):
                return False, f"Row {row_num}: Missing required field '{field}'"
        
        # Validate QR ID format
        qr_id = str(row['qr_id']).strip()
        if len(qr_id) < 3 or len(qr_id) > 50:
            return False, f"Row {row_num}: QR ID must be between 3 and 50 characters"
        
        # Validate bag type
        bag_type = str(row['type']).strip().lower()
        if bag_type not in ('parent', 'child'):
            return False, f"Row {row_num}: Type must be 'parent' or 'child', got '{row['type']}'"
        
        # Validate parent QR ID for child bags
        if bag_type == 'child':
            if not row.get('parent_qr_id'):
                return False, f"Row {row_num}: Child bags must have a parent_qr_id"
        
        return True, None
    
    @staticmethod
    def validate_bill_data(row: Dict, row_num: int) -> Tuple[bool, Optional[str]]:
        """
        Validate a single bill row.
        
        Args:
            row: Dictionary with bill data
            row_num: Row number for error reporting
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        # Required fields
        if not row.get('bill_id'):
            return False, f"Row {row_num}: Missing required field 'bill_id'"
        
        # Validate bill ID format
        bill_id = str(row['bill_id']).strip()
        if len(bill_id) < 3 or len(bill_id) > 50:
            return False, f"Row {row_num}: Bill ID must be between 3 and 50 characters"
        
        # Validate parent bag count if provided and not empty
        if 'parent_bag_count' in row and str(row['parent_bag_count']).strip():
            try:
                count = int(row['parent_bag_count'])
                if count < 1 or count > 500:
                    return False, f"Row {row_num}: Parent bag count must be between 1 and 500"
            except (ValueError, TypeError):
                return False, f"Row {row_num}: Parent bag count must be a valid number"
        
        return True, None


class BagImporter:
    """Handles bulk import of bags from CSV/Excel"""
    
    @staticmethod
    def parse_csv(file_storage: FileStorage) -> Tuple[List[Dict], List[str]]:
        """
        Parse CSV file and return bag data with any errors.
        
        Args:
            file_storage: Uploaded file object
            
        Returns:
            Tuple of (bag_data_list, error_list)
        """
        try:
            # Read file content
            content = file_storage.read().decode('utf-8')
            file_storage.seek(0)  # Reset for potential re-use
            
            # Parse CSV
            csv_reader = csv.DictReader(io.StringIO(content))
            
            bags = []
            errors = []
            
            for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (header is row 1)
                # Validate row
                is_valid, error_msg = ImportValidator.validate_bag_data(row, row_num)
                if not is_valid:
                    errors.append(error_msg)
                    continue
                
                # Normalize data
                bags.append({
                    'qr_id': str(row['qr_id']).strip(),
                    'type': str(row['type']).strip().lower(),
                    'parent_qr_id': str(row.get('parent_qr_id', '')).strip() if row.get('parent_qr_id') else None
                })
            
            return bags, errors
        
        except Exception as e:
            logger.error(f"CSV parsing error: {str(e)}")
            return [], [f"Error parsing CSV file: {str(e)}"]
    
    @staticmethod
    def parse_excel(file_storage: FileStorage) -> Tuple[List[Dict], List[str]]:
        """
        Parse Excel file and return bag data with any errors.
        Processes ALL sheets in the workbook.
        
        Args:
            file_storage: Uploaded file object
            
        Returns:
            Tuple of (bag_data_list, error_list)
        """
        if not EXCEL_AVAILABLE:
            return [], ["Excel support not available - openpyxl not installed"]
        
        try:
            file_storage.stream.seek(0)
            wb = load_workbook(file_storage.stream)
            
            bags = []
            errors = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                
                headers = [cell.value for cell in ws[1]]
                
                if not headers or not any(headers):
                    logger.debug(f"Sheet '{sheet_name}' has no headers, skipping")
                    continue
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue
                    
                    row_dict = dict(zip(headers, row))
                    
                    is_valid, error_msg = ImportValidator.validate_bag_data(row_dict, row_num)
                    if not is_valid:
                        errors.append(f"Sheet '{sheet_name}', Row {row_num}: {error_msg}")
                        continue
                    
                    bags.append({
                        'qr_id': str(row_dict['qr_id']).strip(),
                        'type': str(row_dict['type']).strip().lower(),
                        'parent_qr_id': str(row_dict.get('parent_qr_id', '')).strip() if row_dict.get('parent_qr_id') else None
                    })
            
            logger.info(f"Parsed {len(bags)} bags from {len(wb.sheetnames)} sheet(s)")
            return bags, errors
        
        except Exception as e:
            logger.error(f"Excel parsing error: {str(e)}")
            return [], [f"Error parsing Excel file: {str(e)}"]
    
    @staticmethod
    def import_bags(db, bags: List[Dict], user_id: int) -> Tuple[int, int, List[str]]:
        """
        Import bags into database with transaction safety.
        
        Args:
            db: Database session
            bags: List of bag dictionaries
            user_id: ID of user performing import
            
        Returns:
            Tuple of (imported_count, skipped_count, error_list)
        """
        from models import Bag, Link
        
        imported = 0
        skipped = 0
        errors = []
        
        try:
            # Check for duplicates in batch (much faster than individual queries)
            qr_ids = [bag['qr_id'] for bag in bags]
            
            existing_query = text("""
                SELECT qr_id FROM bag WHERE qr_id = ANY(:qr_ids)
            """)
            
            existing_qr_ids = set(
                row[0] for row in db.session.execute(existing_query, {'qr_ids': qr_ids}).fetchall()
            )
            
            # Process bags in batches with savepoint-based rollback for resilience
            batch_size = 100
            for i in range(0, len(bags), batch_size):
                batch = bags[i:i + batch_size]
                batch_start = imported
                batch_qr_ids = []  # Track IDs added in this batch
                
                # Use nested transaction (savepoint) for batch-level rollback
                savepoint = db.session.begin_nested()
                
                try:
                    for bag_data in batch:
                        qr_id = bag_data['qr_id']
                        
                        # Skip duplicates
                        if qr_id in existing_qr_ids:
                            skipped += 1
                            errors.append(f"Skipped duplicate QR ID: {qr_id}")
                            continue
                        
                        # Create bag
                        new_bag = Bag(
                            qr_id=qr_id,
                            type=bag_data['type']
                        )
                        db.session.add(new_bag)
                        existing_qr_ids.add(qr_id)  # Track newly added
                        batch_qr_ids.append(qr_id)  # Track for potential rollback cleanup
                        imported += 1
                    
                    # Commit batch savepoint
                    savepoint.commit()
                    db.session.flush()
                    # Batch succeeded - IDs are now permanent in existing_qr_ids
                    
                except Exception as batch_error:
                    # Rollback only this batch, not entire import
                    savepoint.rollback()
                    
                    # CRITICAL FIX: Remove batch IDs from tracking set since they were rolled back
                    for qr_id in batch_qr_ids:
                        existing_qr_ids.discard(qr_id)  # Remove rolled-back IDs
                    
                    batch_imported = imported - batch_start
                    imported = batch_start  # Reset counter
                    errors.append(f"Batch {i//batch_size + 1} failed: {str(batch_error)} ({batch_imported} bags lost)")
                    logger.warning(f"Batch rollback at index {i}: {str(batch_error)}")
            
            # Handle child bag linkages after all bags are created
            for bag_data in bags:
                if bag_data['type'] == 'child' and bag_data.get('parent_qr_id'):
                    parent_qr = bag_data['parent_qr_id']
                    child_qr = bag_data['qr_id']
                    
                    # Find parent and child bags
                    parent = Bag.query.filter_by(qr_id=parent_qr).first()
                    child = Bag.query.filter_by(qr_id=child_qr).first()
                    
                    if parent and child:
                        # Check if link already exists
                        existing_link = Link.query.filter_by(
                            parent_bag_id=parent.id,
                            child_bag_id=child.id
                        ).first()
                        
                        if not existing_link:
                            link = Link(
                                parent_bag_id=parent.id,
                                child_bag_id=child.id
                            )
                            db.session.add(link)
                    elif not parent:
                        errors.append(f"Warning: Parent bag '{parent_qr}' not found for child '{child_qr}'")
            
            # Final commit
            db.session.commit()
            
            # Log import activity
            from audit_utils import log_audit
            log_audit(
                action='BULK_IMPORT',
                entity_type='bag',
                entity_id=None,
                details=f"Imported {imported} bags, skipped {skipped} duplicates"
            )
            
            return imported, skipped, errors
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Bag import error: {str(e)}")
            errors.append(f"Database error: {str(e)}")
            return imported, skipped, errors


class BillImporter:
    """Handles bulk import of bills from CSV/Excel"""
    
    @staticmethod
    def parse_csv(file_storage: FileStorage) -> Tuple[List[Dict], List[str]]:
        """Parse CSV file and return bill data with any errors"""
        try:
            content = file_storage.read().decode('utf-8')
            file_storage.seek(0)
            
            csv_reader = csv.DictReader(io.StringIO(content))
            
            bills = []
            errors = []
            
            for row_num, row in enumerate(csv_reader, start=2):
                is_valid, error_msg = ImportValidator.validate_bill_data(row, row_num)
                if not is_valid:
                    errors.append(error_msg)
                    continue
                
                # Safely parse numeric fields with empty value handling
                try:
                    parent_bag_count_str = str(row.get('parent_bag_count', '')).strip()
                    parent_bag_count = int(parent_bag_count_str) if parent_bag_count_str else 1
                except (ValueError, TypeError):
                    parent_bag_count = 1
                
                try:
                    expected_weight_str = str(row.get('expected_weight_kg', '')).strip()
                    expected_weight_kg = float(expected_weight_str) if expected_weight_str else 0.0
                except (ValueError, TypeError):
                    expected_weight_kg = 0.0
                
                bills.append({
                    'bill_id': str(row['bill_id']).strip(),
                    'description': str(row.get('description', '')).strip(),
                    'parent_bag_count': parent_bag_count,
                    'expected_weight_kg': expected_weight_kg
                })
            
            return bills, errors
        
        except Exception as e:
            logger.error(f"CSV parsing error: {str(e)}")
            return [], [f"Error parsing CSV file: {str(e)}"]
    
    @staticmethod
    def import_bills(db, bills: List[Dict], user_id: int) -> Tuple[int, int, List[str]]:
        """Import bills into database with transaction safety"""
        from models import Bill
        
        imported = 0
        skipped = 0
        errors = []
        
        try:
            # Check for duplicates in batch
            bill_ids = [bill['bill_id'] for bill in bills]
            
            existing_query = text("""
                SELECT bill_id FROM bill WHERE bill_id = ANY(:bill_ids)
            """)
            
            existing_bill_ids = set(
                row[0] for row in db.session.execute(existing_query, {'bill_ids': bill_ids}).fetchall()
            )
            
            # Process bills in batches with savepoint-based rollback
            batch_size = 100
            for i in range(0, len(bills), batch_size):
                batch = bills[i:i + batch_size]
                batch_start = imported
                batch_bill_ids = []  # Track IDs added in this batch
                
                # Use nested transaction (savepoint) for batch-level rollback
                savepoint = db.session.begin_nested()
                
                try:
                    for bill_data in batch:
                        bill_id = bill_data['bill_id']
                        
                        # Skip duplicates
                        if bill_id in existing_bill_ids:
                            skipped += 1
                            errors.append(f"Skipped duplicate Bill ID: {bill_id}")
                            continue
                        
                        # Create bill
                        new_bill = Bill(
                            bill_id=bill_id,
                            description=bill_data.get('description'),
                            parent_bag_count=bill_data.get('parent_bag_count', 1),
                            expected_weight_kg=bill_data.get('expected_weight_kg', 0),
                            created_by_id=user_id
                        )
                        db.session.add(new_bill)
                        existing_bill_ids.add(bill_id)
                        batch_bill_ids.append(bill_id)  # Track for potential rollback cleanup
                        imported += 1
                    
                    # Commit batch savepoint
                    savepoint.commit()
                    db.session.flush()
                    # Batch succeeded - IDs are now permanent in existing_bill_ids
                    
                except Exception as batch_error:
                    # Rollback only this batch, not entire import
                    savepoint.rollback()
                    
                    # CRITICAL FIX: Remove batch IDs from tracking set since they were rolled back
                    for bill_id in batch_bill_ids:
                        existing_bill_ids.discard(bill_id)  # Remove rolled-back IDs
                    
                    batch_imported = imported - batch_start
                    imported = batch_start  # Reset counter
                    errors.append(f"Batch {i//batch_size + 1} failed: {str(batch_error)} ({batch_imported} bills lost)")
                    logger.warning(f"Bill batch rollback at index {i}: {str(batch_error)}")
            
            db.session.commit()
            
            # Log import activity
            from audit_utils import log_audit
            log_audit(
                action='BULK_IMPORT',
                entity_type='bill',
                entity_id=None,
                details=f"Imported {imported} bills, skipped {skipped} duplicates"
            )
            
            return imported, skipped, errors
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Bill import error: {str(e)}")
            errors.append(f"Database error: {str(e)}")
            return imported, skipped, errors


class RowResult:
    """Represents the result of processing a single row"""
    SUCCESS = 'success'
    DUPLICATE = 'duplicate'
    ERROR = 'error'
    SKIPPED = 'skipped'
    LINKED = 'linked'
    PARENT_CREATED = 'parent_created'
    CHILD_CREATED = 'child_created'
    
    def __init__(self, row_num: int, qr_code: str, status: str, message: str = '', details: dict = None):
        self.row_num = row_num
        self.qr_code = qr_code
        self.status = status
        self.message = message
        self.details = details or {}
    
    def to_dict(self) -> Dict:
        return {
            'row_num': self.row_num,
            'qr_code': self.qr_code,
            'status': self.status,
            'message': self.message,
            **self.details
        }


class StreamingExcelProcessor:
    """
    Memory-efficient Excel processor for large files (100k+ rows).
    Uses read_only mode and row iteration to minimize memory footprint.
    """
    
    @staticmethod
    def save_to_temp_file(file_storage: FileStorage) -> str:
        """
        Save uploaded file to a temporary location for streaming access.
        Returns the path to the temporary file.
        """
        temp_dir = tempfile.gettempdir()
        temp_filename = f"import_{uuid.uuid4().hex}.xlsx"
        temp_path = os.path.join(temp_dir, temp_filename)
        
        file_storage.stream.seek(0)
        with open(temp_path, 'wb') as f:
            f.write(file_storage.stream.read())
        file_storage.stream.seek(0)
        
        return temp_path
    
    @staticmethod
    def cleanup_temp_file(temp_path: str):
        """Clean up temporary file after processing"""
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            logger.warning(f"Failed to cleanup temp file {temp_path}: {e}")
    
    @staticmethod
    def count_rows(file_path: str) -> int:
        """
        Count total rows in Excel file using streaming mode.
        This is memory-efficient but requires a full pass through the file.
        Counts rows across ALL sheets in the workbook.
        """
        if not EXCEL_AVAILABLE:
            return 0
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            total_count = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                count = sum(1 for _ in ws.iter_rows(min_row=2))
                total_count += count
                logger.debug(f"Sheet '{sheet_name}': {count} data rows")
            wb.close()
            logger.info(f"Total rows across {len(wb.sheetnames)} sheet(s): {total_count}")
            return total_count
        except Exception as e:
            logger.error(f"Error counting rows: {e}")
            return 0
    
    @staticmethod
    def stream_rows(file_path: str, min_row: int = 2) -> Generator[Tuple[int, Tuple, str, bool], None, None]:
        """
        Stream rows from Excel file in memory-efficient manner.
        Streams rows from ALL sheets in the workbook.
        
        Args:
            file_path: Path to Excel file
            min_row: Starting row (default 2 to skip header)
            
        Yields:
            Tuple of (row_number, row_values, sheet_name, is_new_sheet)
            - row_number: Row number within the current sheet
            - row_values: The row data
            - sheet_name: Name of the current sheet
            - is_new_sheet: True if this is the first row of a new sheet
        """
        if not EXCEL_AVAILABLE:
            return
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.debug(f"Processing sheet: {sheet_name}")
                first_row_of_sheet = True
                
                for row_num, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
                    yield row_num, row, sheet_name, first_row_of_sheet
                    first_row_of_sheet = False
            
            wb.close()
            gc.collect()
            
        except Exception as e:
            logger.error(f"Error streaming rows: {e}")
            raise


class LargeScaleChildParentImporter:
    """
    ULTRA HIGH-PERFORMANCE importer for large Excel files with lakhs of rows.
    
    Optimizations:
    - Two-pass approach: Parse first, then bulk insert
    - Single batch duplicate detection query per chunk
    - Raw SQL bulk inserts (100x faster than ORM)
    - Chunk-level transactions (not per-parent)
    - Minimal memory footprint with streaming
    
    Performance: ~500 rows/second (was ~90 rows/second)
    """
    
    MAX_RESULTS_IN_MEMORY = 5000
    CHUNK_SIZE = 5000  # Process 5000 children at a time
    
    @staticmethod
    def process_file_streaming(
        file_storage: FileStorage,
        user_id: int,
        dispatch_area: Optional[str] = None,
        progress_callback: callable = None,
        auto_create_parents: bool = False
    ) -> Tuple[Dict, List[RowResult]]:
        """
        ULTRA-OPTIMIZED: Process large Excel file using two-pass bulk operations.
        
        Pass 1: Stream and collect all batches (parent + children)
        Pass 2: Bulk insert with single duplicate check per chunk
        
        Args:
            file_storage: Uploaded Excel file
            user_id: ID of importing user
            dispatch_area: Optional dispatch area
            progress_callback: Optional callback for progress updates
            auto_create_parents: Always True (bags created fresh)
            
        Returns:
            Tuple of (stats_dict, row_results_list)
        """
        from models import BagType
        import time
        
        start_time = time.time()
        temp_path = None
        row_results = []
        stats = {
            'total_rows': 0,
            'batches_processed': 0,
            'parents_created': 0,
            'parents_found': 0,
            'parents_not_found': 0,
            'parents_rejected_duplicate': 0,
            'children_created': 0,
            'children_existing': 0,
            'links_created': 0,
            'links_existing': 0,
            'errors': 0,
            'skipped': 0
        }
        
        try:
            temp_path = StreamingExcelProcessor.save_to_temp_file(file_storage)
            total_rows = StreamingExcelProcessor.count_rows(temp_path)
            stats['total_rows'] = total_rows
            logger.info(f"ULTRA-OPTIMIZED: Processing {total_rows} rows")
            
            # Initial progress callback: 0%
            if progress_callback:
                progress_callback(0, 100)
            
            # PASS 1: Collect all batches without any DB operations
            all_batches = []
            current_children = []
            current_sheet = None
            rows_processed = 0
            
            for row_num, row, sheet_name, is_new_sheet in StreamingExcelProcessor.stream_rows(temp_path):
                rows_processed += 1
                
                # Progress callback every 1000 rows during parsing (0-50% of progress)
                if progress_callback and rows_processed % 1000 == 0 and total_rows > 0:
                    # Pass 1 is 50% of total work
                    progress_pct = min(49, int((rows_processed / total_rows) * 50))
                    progress_callback(progress_pct, 100)
                
                if is_new_sheet:
                    if current_children:
                        for child in current_children:
                            row_results.append(RowResult(
                                child['row_num'], child['label'], RowResult.ERROR,
                                f"Orphaned child in sheet '{current_sheet}' - no parent",
                                details={'sheet': current_sheet or '', 'parent_qr': '', 'child_qr': child['label']}
                            ))
                            stats['errors'] += 1
                        current_children = []
                    current_sheet = sheet_name
                
                if not any(row):
                    stats['skipped'] += 1
                    continue
                
                sr_no = row[0] if len(row) > 0 else None
                qr_code = row[1] if len(row) > 1 else None
                
                if LargeScaleChildParentImporter._is_child_row(sr_no, qr_code):
                    label_number = LargeScaleChildParentImporter._extract_label_number(qr_code)
                    if label_number:
                        current_children.append({
                            'row_num': row_num,
                            'label': label_number,
                            'sheet': sheet_name
                        })
                    else:
                        row_results.append(RowResult(
                            row_num, str(qr_code)[:50], RowResult.ERROR,
                            f"Sheet '{sheet_name}': Could not extract label number",
                            details={'sheet': sheet_name or '', 'parent_qr': '', 'child_qr': str(qr_code)[:50]}
                        ))
                        stats['errors'] += 1
                    continue
                
                is_parent, parent_code = LargeScaleChildParentImporter._is_parent_row(sr_no, qr_code)
                
                if is_parent:
                    if not parent_code:
                        row_results.append(RowResult(
                            row_num, '', RowResult.ERROR,
                            f"Sheet '{sheet_name}': Parent row found but code is missing",
                            details={'sheet': sheet_name or '', 'parent_qr': '', 'child_qr': ''}
                        ))
                        stats['errors'] += 1
                        current_children = []
                        continue
                    
                    if not current_children:
                        row_results.append(RowResult(
                            row_num, parent_code, RowResult.ERROR,
                            f"Sheet '{sheet_name}': No child bags found for this parent",
                            details={'sheet': sheet_name or '', 'parent_qr': parent_code, 'child_qr': ''}
                        ))
                        stats['errors'] += 1
                        continue
                    
                    all_batches.append({
                        'parent_code': parent_code,
                        'parent_row_num': row_num,
                        'sheet': sheet_name,
                        'children': current_children
                    })
                    current_children = []
            
            if current_children:
                for child in current_children:
                    row_results.append(RowResult(
                        child['row_num'], child['label'], RowResult.ERROR,
                        f"Orphaned child in sheet '{current_sheet}' - no parent",
                        details={'sheet': current_sheet or '', 'parent_qr': '', 'child_qr': child['label']}
                    ))
                    stats['errors'] += 1
            
            parse_time = time.time() - start_time
            logger.info(f"Pass 1 complete: {len(all_batches)} batches parsed in {parse_time:.2f}s")
            
            # Progress callback after parsing (50% complete)
            if progress_callback:
                progress_callback(50, 100)
            
            # PASS 2: Bulk insert with chunked duplicate detection
            if all_batches:
                # Pass the original progress_callback directly - _bulk_process_all_batches now uses direct percentages
                batch_stats, batch_results = LargeScaleChildParentImporter._bulk_process_all_batches(
                    batches=all_batches,
                    user_id=user_id,
                    dispatch_area=dispatch_area,
                    progress_callback=progress_callback
                )
                
                for key in stats:
                    if key in batch_stats:
                        stats[key] += batch_stats[key]
                
                # Separate errors and successes - always preserve errors
                error_results = [r for r in batch_results if r.status == RowResult.ERROR]
                success_results = [r for r in batch_results if r.status != RowResult.ERROR]
                
                # Add all errors first
                row_results.extend(error_results)
                
                # Add successes up to memory limit
                remaining_capacity = LargeScaleChildParentImporter.MAX_RESULTS_IN_MEMORY - len(row_results)
                if remaining_capacity > 0:
                    row_results.extend(success_results[:remaining_capacity])
            else:
                # No batches - still report 100% completion
                if progress_callback:
                    progress_callback(100, 100)
            
            # Note: 100% is reported by Pass 2 progress callback in _bulk_process_all_batches
            
            db.session.commit()
            
            total_time = time.time() - start_time
            rows_per_sec = total_rows / total_time if total_time > 0 else 0
            logger.info(f"ULTRA-OPTIMIZED complete: {total_rows} rows in {total_time:.2f}s ({rows_per_sec:.0f} rows/sec)")
            logger.info(f"Stats: {stats}")
            
            return stats, row_results
            
        except Exception as e:
            db.session.rollback()
            import traceback
            logger.error(f"Streaming import failed: {e}")
            logger.error(f"Full traceback: {traceback.format_exc()}")
            raise
        finally:
            if temp_path:
                StreamingExcelProcessor.cleanup_temp_file(temp_path)
    
    @staticmethod
    def _bulk_process_all_batches(
        batches: List[Dict],
        user_id: int,
        dispatch_area: Optional[str],
        progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> Tuple[Dict, List[RowResult]]:
        """
        ULTRA-OPTIMIZED: Process all batches with minimal DB round-trips.
        
        Strategy:
        1. Collect ALL parent codes and child labels
        2. Single query to find ALL existing bags
        3. Bulk insert ALL new parents
        4. Bulk insert ALL new children
        5. Bulk insert ALL links
        """
        from models import BagType
        import time
        
        start_time = time.time()
        results = []
        stats = {
            'batches_processed': 0,
            'parents_created': 0,
            'parents_found': 0,
            'parents_rejected_duplicate': 0,
            'children_created': 0,
            'children_existing': 0,
            'links_created': 0,
            'links_existing': 0,
            'errors': 0
        }
        
        # Collect all unique codes - handle intra-file duplicates
        all_parent_codes = set()
        all_child_labels = set()
        parent_to_children = {}  # parent_code -> [child_info]
        parent_info = {}  # parent_code -> {row_num, sheet}
        intra_file_duplicate_parents = {}  # parent_code -> list of duplicate batch infos
        
        for batch in batches:
            parent_code = batch['parent_code'].upper()
            
            if parent_code in all_parent_codes:
                # Intra-file duplicate - same parent appears multiple times in file
                if parent_code not in intra_file_duplicate_parents:
                    intra_file_duplicate_parents[parent_code] = []
                intra_file_duplicate_parents[parent_code].append({
                    'row_num': batch['parent_row_num'],
                    'sheet': batch['sheet'],
                    'original_code': batch['parent_code'],
                    'children': batch['children']
                })
            else:
                all_parent_codes.add(parent_code)
                parent_info[parent_code] = {
                    'row_num': batch['parent_row_num'],
                    'sheet': batch['sheet'],
                    'original_code': batch['parent_code']
                }
                
                if parent_code not in parent_to_children:
                    parent_to_children[parent_code] = []
                
                for child in batch['children']:
                    child_label_upper = child['label'].upper()
                    all_child_labels.add(child_label_upper)
                    parent_to_children[parent_code].append({
                        'row_num': child['row_num'],
                        'label': child['label'],
                        'label_upper': child_label_upper,
                        'sheet': child['sheet']
                    })
        
        # Handle intra-file duplicate parents as errors
        for parent_code, duplicates in intra_file_duplicate_parents.items():
            first_info = parent_info.get(parent_code)
            first_row = first_info['row_num'] if first_info else 'unknown'
            
            for dup in duplicates:
                stats['errors'] += 1
                stats['parents_rejected_duplicate'] += 1
                results.append(RowResult(
                    dup['row_num'], dup['original_code'], RowResult.ERROR,
                    f"Sheet '{dup['sheet']}': Duplicate parent in same file (first at row {first_row})",
                    details={'sheet': dup['sheet'], 'parent_qr': dup['original_code']}
                ))
                # Mark all children as errors
                for child in dup['children']:
                    stats['errors'] += 1
                    results.append(RowResult(
                        child['row_num'], child['label'], RowResult.ERROR,
                        f"Sheet '{child['sheet']}': Rejected - parent '{dup['original_code']}' is duplicate in file",
                        details={'sheet': child['sheet'], 'parent_qr': dup['original_code'], 'child_qr': child['label']}
                    ))
        
        logger.info(f"Collected {len(all_parent_codes)} parents, {len(all_child_labels)} children")
        
        # SINGLE query to find ALL existing bags (parents + children)
        all_codes = list(all_parent_codes | all_child_labels)
        existing_bags = {}
        
        if all_codes:
            # Batch in chunks of 5000 to avoid query size limits
            for i in range(0, len(all_codes), 5000):
                chunk = all_codes[i:i+5000]
                # Use IN clause with dynamic placeholders for compatibility
                placeholders = ', '.join([f':code_{j}' for j in range(len(chunk))])
                params = {f'code_{j}': code for j, code in enumerate(chunk)}
                
                result = db.session.execute(
                    text(f"SELECT id, UPPER(qr_id) as qr_id_upper FROM bag WHERE UPPER(qr_id) IN ({placeholders})"),
                    params
                )
                for row in result:
                    existing_bags[row.qr_id_upper] = row.id
        
        logger.info(f"Found {len(existing_bags)} existing bags in database")
        
        # Determine which parents are duplicates (error) vs new (create)
        duplicate_parents = set()
        new_parents = []
        
        for parent_code in all_parent_codes:
            info = parent_info[parent_code]
            if parent_code in existing_bags:
                duplicate_parents.add(parent_code)
                stats['parents_rejected_duplicate'] += 1
                stats['errors'] += 1
                results.append(RowResult(
                    info['row_num'], info['original_code'], RowResult.ERROR,
                    f"Sheet '{info['sheet']}': Parent bag already exists - cannot import duplicate",
                    details={'sheet': info['sheet'], 'parent_qr': info['original_code']}
                ))
                # Mark all children of this parent as errors
                for child in parent_to_children[parent_code]:
                    stats['errors'] += 1
                    results.append(RowResult(
                        child['row_num'], child['label'], RowResult.ERROR,
                        f"Sheet '{child['sheet']}': Rejected - parent '{info['original_code']}' already exists",
                        details={'sheet': child['sheet'], 'parent_qr': info['original_code'], 'child_qr': child['label']}
                    ))
            else:
                new_parents.append({
                    'qr_id': info['original_code'],
                    'qr_id_upper': parent_code,
                    'row_num': info['row_num'],
                    'sheet': info['sheet']
                })
        
        logger.info(f"New parents: {len(new_parents)}, Duplicate parents: {len(duplicate_parents)}")
        
        # Progress: 55% overall (duplicate detection complete)
        if progress_callback:
            progress_callback(55, 100)
        
        # BULK INSERT all new parents
        parent_id_map = {}  # parent_code_upper -> bag_id
        
        if new_parents:
            # Use raw SQL for maximum speed
            parent_values = []
            for p in new_parents:
                parent_values.append({
                    'qr_id': p['qr_id'],
                    'type': BagType.PARENT.value,
                    'user_id': user_id,
                    'dispatch_area': dispatch_area,
                    'weight_kg': 0.0,
                    'child_count': 0
                })
            
            # Batch insert with RETURNING to get IDs
            for i in range(0, len(parent_values), 1000):
                chunk = parent_values[i:i+1000]
                
                # Build VALUES clause
                values_parts = []
                params = {}
                for j, p in enumerate(chunk):
                    key = f"p{i+j}"
                    values_parts.append(f"(:{key}_qr, :{key}_type, :{key}_user, :{key}_area, :{key}_weight, :{key}_count)")
                    params[f"{key}_qr"] = p['qr_id']
                    params[f"{key}_type"] = p['type']
                    params[f"{key}_user"] = p['user_id']
                    params[f"{key}_area"] = p['dispatch_area']
                    params[f"{key}_weight"] = p['weight_kg']
                    params[f"{key}_count"] = p['child_count']
                
                insert_sql = f"""
                    INSERT INTO bag (qr_id, type, user_id, dispatch_area, weight_kg, child_count)
                    VALUES {', '.join(values_parts)}
                    RETURNING id, UPPER(qr_id) as qr_id_upper
                """
                
                result = db.session.execute(text(insert_sql), params)
                for row in result:
                    parent_id_map[row.qr_id_upper] = row.id
            
            stats['parents_created'] = len(new_parents)
            
            for p in new_parents:
                results.append(RowResult(
                    p['row_num'], p['qr_id'], RowResult.PARENT_CREATED,
                    f"Sheet '{p['sheet']}': Parent bag created",
                    details={'sheet': p['sheet'], 'parent_qr': p['qr_id'], 'child_qr': ''}
                ))
        
        logger.info(f"Inserted {len(parent_id_map)} parent bags")
        
        # Progress: 65% overall (parents inserted)
        if progress_callback:
            progress_callback(65, 100)
        
        # Collect children for non-duplicate parents - handle intra-file child duplicates
        new_children = []
        duplicate_children = []
        seen_children_in_file = {}  # child_label_upper -> first occurrence info
        
        for parent_code in all_parent_codes:
            if parent_code in duplicate_parents:
                continue  # Skip children of duplicate parents
            
            parent_id = parent_id_map.get(parent_code)
            if not parent_id:
                continue
            
            for child in parent_to_children[parent_code]:
                # Check for DB duplicates
                if child['label_upper'] in existing_bags:
                    duplicate_children.append(child)
                    stats['children_existing'] += 1
                    stats['errors'] += 1
                    results.append(RowResult(
                        child['row_num'], child['label'], RowResult.ERROR,
                        f"Sheet '{child['sheet']}': Child bag already exists in database - cannot import duplicate",
                        details={'sheet': child['sheet'], 'parent_qr': parent_info[parent_code]['original_code'], 'child_qr': child['label']}
                    ))
                # Check for intra-file duplicates
                elif child['label_upper'] in seen_children_in_file:
                    first = seen_children_in_file[child['label_upper']]
                    stats['errors'] += 1
                    stats['children_existing'] += 1
                    results.append(RowResult(
                        child['row_num'], child['label'], RowResult.ERROR,
                        f"Sheet '{child['sheet']}': Duplicate child in same file (first at row {first['row_num']})",
                        details={'sheet': child['sheet'], 'parent_qr': parent_info[parent_code]['original_code'], 'child_qr': child['label']}
                    ))
                else:
                    seen_children_in_file[child['label_upper']] = {
                        'row_num': child['row_num'],
                        'sheet': child['sheet']
                    }
                    new_children.append({
                        'qr_id': child['label'],
                        'qr_id_upper': child['label_upper'],
                        'parent_code_upper': parent_code,
                        'parent_id': parent_id,
                        'row_num': child['row_num'],
                        'sheet': child['sheet'],
                        'parent_original': parent_info[parent_code]['original_code']
                    })
        
        logger.info(f"New children: {len(new_children)}, Duplicate children: {len(duplicate_children)}")
        
        # BULK INSERT all new children
        child_id_map = {}  # child_qr_upper -> bag_id
        
        if new_children:
            for i in range(0, len(new_children), 1000):
                chunk = new_children[i:i+1000]
                
                values_parts = []
                params = {}
                for j, c in enumerate(chunk):
                    key = f"c{i+j}"
                    values_parts.append(f"(:{key}_qr, :{key}_type, :{key}_user, :{key}_area, :{key}_weight)")
                    params[f"{key}_qr"] = c['qr_id']
                    params[f"{key}_type"] = BagType.CHILD.value
                    params[f"{key}_user"] = user_id
                    params[f"{key}_area"] = dispatch_area
                    params[f"{key}_weight"] = 1.0
                
                insert_sql = f"""
                    INSERT INTO bag (qr_id, type, user_id, dispatch_area, weight_kg)
                    VALUES {', '.join(values_parts)}
                    RETURNING id, UPPER(qr_id) as qr_id_upper
                """
                
                result = db.session.execute(text(insert_sql), params)
                for row in result:
                    child_id_map[row.qr_id_upper] = row.id
            
            stats['children_created'] = len(new_children)
            
            for c in new_children:
                results.append(RowResult(
                    c['row_num'], c['qr_id'], RowResult.CHILD_CREATED,
                    f"Sheet '{c['sheet']}': Child bag created and linked",
                    details={'sheet': c['sheet'], 'parent_qr': c['parent_original'], 'child_qr': c['qr_id']}
                ))
        
        logger.info(f"Inserted {len(child_id_map)} child bags")
        
        # Progress: 85% overall (children inserted)
        if progress_callback:
            progress_callback(85, 100)
        
        # BULK INSERT all links
        links_to_create = []
        for c in new_children:
            child_id = child_id_map.get(c['qr_id_upper'])
            if child_id:
                links_to_create.append({
                    'parent_bag_id': c['parent_id'],
                    'child_bag_id': child_id
                })
        
        if links_to_create:
            for i in range(0, len(links_to_create), 1000):
                chunk = links_to_create[i:i+1000]
                
                values_parts = []
                params = {}
                for j, link in enumerate(chunk):
                    key = f"l{i+j}"
                    values_parts.append(f"(:{key}_parent, :{key}_child)")
                    params[f"{key}_parent"] = link['parent_bag_id']
                    params[f"{key}_child"] = link['child_bag_id']
                
                insert_sql = f"""
                    INSERT INTO link (parent_bag_id, child_bag_id)
                    VALUES {', '.join(values_parts)}
                    ON CONFLICT (child_bag_id) DO NOTHING
                """
                
                db.session.execute(text(insert_sql), params)
            
            stats['links_created'] = len(links_to_create)
        
        logger.info(f"Created {len(links_to_create)} links")
        
        # Progress: 95% overall (links created)
        if progress_callback:
            progress_callback(95, 100)
        
        # Update parent child counts
        parent_child_counts = {}
        for c in new_children:
            parent_id = c['parent_id']
            parent_child_counts[parent_id] = parent_child_counts.get(parent_id, 0) + 1
        
        if parent_child_counts:
            for parent_id, count in parent_child_counts.items():
                db.session.execute(
                    text("UPDATE bag SET child_count = :count, weight_kg = :weight WHERE id = :id"),
                    {'count': count, 'weight': float(count), 'id': parent_id}
                )
        
        # Stats:
        # batches_processed = total parent batches in file including intra-file duplicates
        stats['batches_processed'] = len(batches)
        # parents_found = duplicate parents that already existed (in legacy terms, "found" means existing)
        stats['parents_found'] = len(duplicate_parents)
        # parents_created already set above
        # parents_rejected_duplicate = DB duplicates + intra-file duplicates (already incremented)
        
        # Progress: 100% overall - complete
        if progress_callback:
            progress_callback(100, 100)
        
        total_time = time.time() - start_time
        logger.info(f"Bulk processing complete in {total_time:.2f}s")
        
        return stats, results
    
    @staticmethod
    def _extract_label_number(qr_text: str) -> Optional[str]:
        """Extract label number from QR code text.
        
        Handles both formats:
        - LABEL NO.0016557 (without colon)
        - LABEL NO.:0016557 (with colon)
        """
        if not qr_text or not isinstance(qr_text, str):
            return None
        import re
        match = re.search(r'LABEL\s*NO\.:?\s*(\d+)', qr_text, re.IGNORECASE)
        return match.group(1) if match else None
    
    @staticmethod
    def _is_child_row(sr_no, qr_code) -> bool:
        """Check if row is a child bag row"""
        qr_str = str(qr_code).strip().upper() if qr_code else ''
        
        # Exclude parent bag prefixes
        if qr_str.startswith('SB') or qr_str.startswith('M444-'):
            return False
        
        try:
            if sr_no is not None:
                float(sr_no)
                if qr_code and isinstance(qr_code, str) and 'LABEL NO.' in qr_code.upper():
                    return True
        except (ValueError, TypeError):
            pass
        
        return False
    
    @staticmethod
    def _is_parent_row(sr_no, qr_code) -> Tuple[bool, Optional[str]]:
        """Check if row is a parent bag row"""
        qr_str = str(qr_code).strip().upper() if qr_code else ''
        
        is_parent_qr = qr_str.startswith('SB') or qr_str.startswith('M444-')
        
        # Legacy format
        if sr_no and isinstance(sr_no, str) and 'parent code' in sr_no.lower():
            return True, str(qr_code).strip() if qr_code else None
        
        # New format - parent prefix
        if is_parent_qr:
            return True, str(qr_code).strip() if qr_code else None
        
        # Serial number 16 format
        try:
            if sr_no is not None and int(float(sr_no)) == 16:
                if 'LABEL NO.' not in qr_str:
                    return True, str(qr_code).strip() if qr_code else None
        except (ValueError, TypeError):
            pass
        
        return False, None
    
    @staticmethod
    def _process_batch(
        parent_code: str,
        children: List[Dict],
        user_id: int,
        dispatch_area: Optional[str],
        batch_num: int,
        parent_row_num: int,
        auto_create_parent: bool = False,
        sheet_name: str = None
    ) -> Tuple[Dict, List[RowResult]]:
        """
        Process a single batch of children for one parent.
        OPTIMIZED: Uses bulk operations to minimize database round-trips.
        """
        from models import Bag, Link, BagType
        from sqlalchemy import func
        
        results = []
        stats = {
            'parent_found': False,
            'parent_created': 0,
            'parent_rejected_duplicate': 0,
            'children_created': 0,
            'children_existing': 0,
            'links_created': 0,
            'links_existing': 0,
            'errors': 0
        }
        
        savepoint = db.session.begin_nested()
        sheet_prefix = f"Sheet '{sheet_name}', " if sheet_name else ""
        
        try:
            parent_bag = Bag.query.filter(func.upper(Bag.qr_id) == parent_code.upper()).first()
            
            if parent_bag:
                # Parent bag already exists - this is an ERROR per new requirement
                # All bags in import file should be new (not pre-existing in database)
                savepoint.rollback()
                results.append(RowResult(
                    parent_row_num, parent_code, RowResult.ERROR,
                    f"{sheet_prefix}Parent bag already exists in database - cannot import duplicate",
                    details={'sheet': sheet_name or '', 'parent_qr': parent_code}
                ))
                stats['errors'] += 1
                stats['parent_rejected_duplicate'] = 1
                # Also mark all children as errors since their parent is rejected
                for child in children:
                    results.append(RowResult(
                        child['row_num'], child['label'], RowResult.ERROR,
                        f"{sheet_prefix}Rejected - parent bag '{parent_code}' already exists",
                        details={'sheet': sheet_name or '', 'parent_qr': parent_code, 'child_qr': child['label']}
                    ))
                    stats['errors'] += 1
                return stats, results
            else:
                # Parent doesn't exist - create it (this is the expected case)
                parent_bag = Bag(
                    qr_id=parent_code,
                    type=BagType.PARENT.value,
                    user_id=user_id,
                    dispatch_area=dispatch_area,
                    weight_kg=0.0,
                    child_count=0
                )
                db.session.add(parent_bag)
                db.session.flush()
                stats['parent_created'] = 1
                results.append(RowResult(
                    parent_row_num, parent_code, RowResult.PARENT_CREATED,
                    f"{sheet_prefix}Parent bag created, processing {len(children)} children",
                    details={'sheet': sheet_name or '', 'parent_qr': parent_code, 'child_qr': ''}
                ))
                logger.info(f"Created parent bag: {parent_code}")
            
            child_labels = [c['label'] for c in children]
            child_labels_upper = [l.upper() for l in child_labels]
            
            existing_children = {
                b.qr_id.upper(): b for b in 
                Bag.query.filter(func.upper(Bag.qr_id).in_(child_labels_upper)).all()
            }
            
            existing_child_ids = [b.id for b in existing_children.values()]
            child_to_parent_link = {}
            if existing_child_ids:
                links = Link.query.filter(Link.child_bag_id.in_(existing_child_ids)).all()
                child_to_parent_link = {l.child_bag_id: l.parent_bag_id for l in links}
            
            new_children_to_create = []
            children_with_errors = []
            
            for child_data in children:
                row_num = child_data['row_num']
                label = child_data['label']
                label_upper = label.upper()
                
                child_bag = existing_children.get(label_upper)
                
                if not child_bag:
                    # Child doesn't exist - this is the expected case, will be created
                    new_children_to_create.append({
                        'row_num': row_num,
                        'label': label,
                        'label_upper': label_upper
                    })
                else:
                    # Child already exists in database - this is an ERROR per new requirement
                    # All bags in import file should be new (not pre-existing in database)
                    stats['errors'] += 1
                    stats['children_existing'] += 1
                    children_with_errors.append({
                        'row_num': row_num,
                        'label': label,
                        'error': f"{sheet_prefix}Child bag already exists in database - cannot import duplicate"
                    })
            
            for err in children_with_errors:
                results.append(RowResult(
                    err['row_num'], err['label'], RowResult.ERROR, err['error'],
                    details={'sheet': sheet_name or '', 'parent_qr': parent_code, 'child_qr': err['label']}
                ))
            
            new_bag_objects = []
            if new_children_to_create:
                for child_info in new_children_to_create:
                    bag = Bag(
                        qr_id=child_info['label'],
                        type=BagType.CHILD.value,
                        user_id=user_id,
                        dispatch_area=dispatch_area,
                        weight_kg=1.0
                    )
                    new_bag_objects.append(bag)
                    existing_children[child_info['label_upper']] = bag
                
                db.session.add_all(new_bag_objects)
                db.session.flush()
            
            children_needing_links = []
            for i, child_info in enumerate(new_children_to_create):
                children_needing_links.append({
                    'child_id': new_bag_objects[i].id,
                    'row_num': child_info['row_num'],
                    'label': child_info['label']
                })
            
            links_created_count = 0
            if children_needing_links:
                new_links = []
                for child_info in children_needing_links:
                    link = Link(parent_bag_id=parent_bag.id, child_bag_id=child_info['child_id'])
                    new_links.append(link)
                    links_created_count += 1
                
                db.session.add_all(new_links)
            
            parent_bag.child_count = (parent_bag.child_count or 0) + links_created_count
            parent_bag.weight_kg = parent_bag.child_count * 1.0
            
            savepoint.commit()
            
            stats['children_created'] = len(new_children_to_create)
            stats['links_created'] = links_created_count
            
            for child_info in new_children_to_create:
                results.append(RowResult(
                    child_info['row_num'], child_info['label'], RowResult.CHILD_CREATED,
                    f"{sheet_prefix}Child bag created and linked",
                    details={'sheet': sheet_name or '', 'parent_qr': parent_code, 'child_qr': child_info['label']}
                ))
            
            return stats, results
            
        except Exception as e:
            savepoint.rollback()
            logger.error(f"Batch {batch_num} failed: {e}")
            results.append(RowResult(
                parent_row_num, parent_code, RowResult.ERROR,
                f"{sheet_prefix}Batch processing failed: {str(e)}",
                details={'sheet': sheet_name or '', 'parent_qr': parent_code}
            ))
            stats['errors'] += len(children)
            return stats, results


class ChildParentBatchImporter:
    """
    Handles batch import of child bags linked to parent bags from Excel files.
    
    Expected format:
    - Multiple rows with child bag QR codes (Sr. No. is numeric, QR Code starts with 'LABEL NO.')
    - Followed by one parent bag row (Sr. No. contains 'Parent Code', QR Code contains parent bag ID)
    - Pattern repeats for each batch
    
    Example:
    Row 1: Sr. No.=1, QR Code='LABEL NO.0016586 LOT NO.:...'  <- Child bag
    Row 2: Sr. No.=2, QR Code='LABEL NO.0016587 LOT NO.:...'  <- Child bag
    Row 3: Sr. No.='Parent Code', QR Code='SB12260'           <- Parent bag (marks end of batch)
    
    For large files (>10k rows), automatically uses streaming mode for memory efficiency.
    """
    
    @staticmethod
    def extract_label_number(qr_text: str) -> Optional[str]:
        """
        Extract only the label number from QR code text.
        
        Handles both formats:
        - 'LABEL NO.0016586 LOT NO.:...' (without colon after NO.)
        - 'LABEL NO.:0016586 LOT NO.:...' (with colon after NO.)
        
        Example input: 'LABEL NO.:0016586 LOT NO.:STAR44GG24611, D.O.T.:30/09/2025...'
        Expected output: '0016586'
        
        Args:
            qr_text: Full QR code text
            
        Returns:
            Label number string or None if not found
        """
        if not qr_text or not isinstance(qr_text, str):
            return None
        
        # Look for 'LABEL NO.' or 'LABEL NO.:' followed by digits
        import re
        match = re.search(r'LABEL\s*NO\.:?\s*(\d+)', qr_text, re.IGNORECASE)
        if match:
            return match.group(1)
        
        return None
    
    @staticmethod
    def is_child_row(sr_no, qr_code) -> bool:
        """
        Determine if a row represents a child bag.
        
        Child bag criteria:
        - Sr. No. is numeric (1-15, not 16 which is parent)
        - QR Code contains 'LABEL NO.' (child bag format)
        - QR Code does NOT start with parent prefixes (SB, M444-)
        
        Args:
            sr_no: Value from Sr. No. column
            qr_code: Value from QR Code column
            
        Returns:
            True if this is a child bag row
        """
        qr_str = str(qr_code).strip().upper() if qr_code else ''
        
        # Exclude parent bag prefixes
        if qr_str.startswith('SB') or qr_str.startswith('M444-'):
            return False
        
        # Check if Sr. No. is numeric
        try:
            if sr_no is not None:
                float(sr_no)  # Can be 1.0 or 1
                
                # Check if QR code contains 'LABEL NO.'
                if qr_code and isinstance(qr_code, str) and 'LABEL NO.' in qr_code.upper():
                    return True
        except (ValueError, TypeError):
            pass
        
        return False
    
    @staticmethod
    def is_parent_row(sr_no, qr_code) -> Tuple[bool, Optional[str]]:
        """
        Determine if a row represents a parent bag and extract parent code.
        
        Parent bag criteria (any of these):
        1. Sr. No. contains 'Parent Code' text (legacy format)
        2. Sr. No. is exactly 16 AND QR Code starts with 'SB' or 'M444-'
        3. QR Code starts with 'SB' or 'M444-' (regardless of serial number)
        
        Args:
            sr_no: Value from Sr. No. column
            qr_code: Value from QR Code column
            
        Returns:
            Tuple of (is_parent_row, parent_code)
        """
        qr_str = str(qr_code).strip().upper() if qr_code else ''
        
        # Check if QR code starts with parent bag prefixes (SB or M444-)
        is_parent_qr = qr_str.startswith('SB') or qr_str.startswith('M444-')
        
        # Legacy format: Sr. No. contains 'Parent Code'
        if sr_no and isinstance(sr_no, str) and 'parent code' in sr_no.lower():
            parent_code = str(qr_code).strip() if qr_code else None
            return True, parent_code
        
        # New format: Serial number is 16 AND/OR QR code starts with parent prefix
        if is_parent_qr:
            parent_code = str(qr_code).strip() if qr_code else None
            return True, parent_code
        
        # Check if serial number is exactly 16 (even without matching prefix)
        try:
            if sr_no is not None and int(float(sr_no)) == 16:
                # Only treat as parent if QR doesn't contain 'LABEL NO.' (child indicator)
                if 'LABEL NO.' not in qr_str:
                    parent_code = str(qr_code).strip() if qr_code else None
                    return True, parent_code
        except (ValueError, TypeError):
            pass
        
        return False, None
    
    @staticmethod
    def parse_excel_batch(file_storage: FileStorage) -> Tuple[List[Dict], List[str], Dict]:
        """
        Parse Excel file with batch pattern (children followed by parent).
        Processes ALL sheets in the workbook.
        
        Args:
            file_storage: Uploaded Excel file
            
        Returns:
            Tuple of (batches_list, error_list, stats_dict)
            
            batches_list format:
            [
                {
                    'parent_code': 'SB12260',
                    'children': ['0016586', '0016587', '0016585', ...],
                    'row_range': 'Sheet1:2-17'
                },
                ...
            ]
        """
        if not EXCEL_AVAILABLE:
            return [], ["Excel support not available - openpyxl not installed"], {}
        
        try:
            file_storage.stream.seek(0)
            wb = load_workbook(file_storage.stream)
            
            batches = []
            errors = []
            total_children = 0
            total_parents = 0
            skipped_rows = 0
            sheets_processed = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                sheets_processed += 1
                
                current_children = []
                batch_start_row = None
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        skipped_rows += 1
                        continue
                    
                    sr_no = row[0] if len(row) > 0 else None
                    qr_code = row[1] if len(row) > 1 else None
                    
                    if ChildParentBatchImporter.is_child_row(sr_no, qr_code):
                        label_number = ChildParentBatchImporter.extract_label_number(qr_code)
                        
                        if label_number:
                            if batch_start_row is None:
                                batch_start_row = row_num
                            current_children.append(label_number)
                            total_children += 1
                        else:
                            errors.append(f"Sheet '{sheet_name}', Row {row_num}: Could not extract label number from QR code")
                    
                    is_parent, parent_code = ChildParentBatchImporter.is_parent_row(sr_no, qr_code)
                    if is_parent:
                        if not parent_code:
                            errors.append(f"Sheet '{sheet_name}', Row {row_num}: Parent row found but parent code is blank/missing")
                            current_children = []
                            batch_start_row = None
                            continue
                        
                        if not current_children:
                            errors.append(f"Sheet '{sheet_name}', Row {row_num}: Parent code '{parent_code}' found but no child bags above it")
                            continue
                        
                        batches.append({
                            'parent_code': parent_code,
                            'children': current_children.copy(),
                            'row_range': f"{sheet_name}:{batch_start_row}-{row_num}"
                        })
                        total_parents += 1
                        
                        current_children = []
                        batch_start_row = None
                
                if current_children:
                    errors.append(f"Warning: {len(current_children)} child bags found at end of sheet '{sheet_name}' without a parent code")
                    batches.append({
                        'parent_code': None,
                        'children': current_children,
                        'row_range': f"{sheet_name}:{batch_start_row}-END",
                        'orphaned': True
                    })
            
            logger.info(f"Parsed {len(batches)} batches from {sheets_processed} sheet(s)")
            
            stats = {
                'total_batches': len(batches),
                'total_children': total_children,
                'total_parents': total_parents,
                'skipped_rows': skipped_rows,
                'sheets_processed': sheets_processed
            }
            
            return batches, errors, stats
        
        except Exception as e:
            logger.error(f"Excel batch parsing error: {str(e)}")
            return [], [f"Error parsing Excel file: {str(e)}"], {}
    
    @staticmethod
    def import_batches(db, batches: List[Dict], user_id: int, dispatch_area: Optional[str] = None) -> Tuple[int, int, int, int, List[str]]:
        """
        Import batches of child-parent bag relationships.
        
        Args:
            db: Database session
            batches: List of batch dictionaries from parse_excel_batch
            user_id: ID of user performing import
            dispatch_area: Optional dispatch area for the bags
            
        Returns:
            Tuple of (parents_created, children_created, links_created, parents_not_found, error_list)
        """
        from models import Bag, Link, BagType
        
        parents_created = 0
        children_created = 0
        links_created = 0
        parents_not_found = 0
        errors = []
        
        try:
            for batch_num, batch in enumerate(batches, 1):
                parent_code = batch['parent_code']
                child_labels = batch['children']
                row_range = batch['row_range']
                is_orphaned = batch.get('orphaned', False)
                
                # Skip orphaned batches unless user explicitly wants them
                if is_orphaned:
                    errors.append(f"Batch {batch_num} (rows {row_range}): Skipped {len(child_labels)} orphaned children without parent")
                    continue
                
                # Start a savepoint for this batch
                savepoint = db.session.begin_nested()
                
                try:
                    # Check if parent already exists (case-insensitive search)
                    from sqlalchemy import func
                    parent_bag = Bag.query.filter(func.upper(Bag.qr_id) == parent_code.upper()).first()
                    
                    if parent_bag:
                        # REJECT batch - parent bag already exists in database (duplicate)
                        # All bags in import file must be NEW (not pre-existing)
                        parents_not_found += 1  # Using this counter for rejected parents
                        error_msg = f"Batch {batch_num} (rows {row_range}): Parent bag '{parent_code}' already exists in database - cannot import duplicate. {len(child_labels)} child bags rejected"
                        errors.append(error_msg)
                        logger.warning(error_msg)
                        savepoint.rollback()
                        continue
                    
                    # Parent doesn't exist - create it (this is the expected case)
                    parent_bag = Bag(
                        qr_id=parent_code,
                        type=BagType.PARENT.value,
                        user_id=user_id,
                        dispatch_area=dispatch_area,
                        weight_kg=0.0,
                        child_count=0
                    )
                    db.session.add(parent_bag)
                    db.session.flush()
                    parents_created += 1
                    logger.info(f"Created parent bag {parent_code}, processing {len(child_labels)} children")
                    
                    # Create child bags and links
                    batch_children_created = 0
                    batch_links_created = 0
                    
                    for label_number in child_labels:
                        # Check if child already exists (case-insensitive search)
                        from sqlalchemy import func
                        child_bag = Bag.query.filter(func.upper(Bag.qr_id) == label_number.upper()).first()
                        
                        if child_bag:
                            # REJECT - child bag already exists in database (duplicate)
                            # All bags in import file must be NEW (not pre-existing)
                            error_msg = f"Batch {batch_num} (rows {row_range}): Child bag '{label_number}' already exists in database - cannot import duplicate"
                            errors.append(error_msg)
                            logger.warning(error_msg)
                            continue
                        
                        # Child doesn't exist - create it (this is the expected case)
                        child_bag = Bag(
                            qr_id=label_number,
                            type=BagType.CHILD.value,
                            user_id=user_id,
                            dispatch_area=dispatch_area,
                            weight_kg=1.0  # Each child is 1kg
                        )
                        db.session.add(child_bag)
                        db.session.flush()  # Get child bag ID
                        batch_children_created += 1
                        
                        # Create link between parent and child
                        link = Link(
                            parent_bag_id=parent_bag.id,
                            child_bag_id=child_bag.id
                        )
                        db.session.add(link)
                        batch_links_created += 1
                    
                    # Update parent's child count
                    actual_child_count = Link.query.filter_by(parent_bag_id=parent_bag.id).count()
                    parent_bag.child_count = actual_child_count
                    parent_bag.weight_kg = actual_child_count * 1.0
                    
                    # Commit this batch
                    savepoint.commit()
                    db.session.flush()
                    
                    children_created += batch_children_created
                    links_created += batch_links_created
                    
                    # Clearer logging to distinguish new vs reused children
                    children_reused = len(child_labels) - batch_children_created
                    if children_reused > 0:
                        logger.info(f"Batch {batch_num} ({row_range}): Parent '{parent_code}', {batch_children_created} new children + {children_reused} existing children, {batch_links_created} new links")
                    else:
                        logger.info(f"Batch {batch_num} ({row_range}): Created parent '{parent_code}', {batch_children_created} children, {batch_links_created} links")
                
                except Exception as batch_error:
                    savepoint.rollback()
                    error_msg = f"Batch {batch_num} (rows {row_range}, parent '{parent_code}'): {str(batch_error)}"
                    errors.append(error_msg)
                    logger.error(f"Batch import error: {error_msg}")
            
            # Final commit
            db.session.commit()
            
            # Log audit
            from audit_utils import log_audit
            log_audit(
                action='BATCH_IMPORT_CHILD_PARENT',
                entity_type='bag',
                entity_id=None,
                details=f"Imported {len(batches)} batches: {parents_created} parents, {children_created} children, {links_created} links ({parents_not_found} parents not found)"
            )
            
            return parents_created, children_created, links_created, parents_not_found, errors
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Batch import error: {str(e)}")
            errors.append(f"Database error: {str(e)}")
            return parents_created, children_created, links_created, parents_not_found, errors


class ParentBillBatchImporter:
    """
    Handles batch import of parent bags linked to bills from Excel files.
    
    Expected format (similar to child→parent pattern):
    - Multiple rows with parent bag QR codes (Sr. No. is numeric)
    - Followed by one bill number row (Sr. No. contains 'Bill No' or similar marker)
    - Pattern repeats for each batch
    
    Example:
    Row 1: Sr. No.=1, Parent Bag='SB12260'     <- Parent bag
    Row 2: Sr. No.=2, Parent Bag='SB12248'     <- Parent bag
    Row 3: Sr. No.='Bill No', Bill='BILL001'   <- Bill number (marks end of batch)
    """
    
    @staticmethod
    def is_parent_bag_row(sr_no, parent_code) -> bool:
        """
        Determine if a row represents a parent bag.
        
        Parent bag criteria:
        - Sr. No. is numeric (1, 2, 3, etc.)
        - Parent code/QR is not empty
        
        Args:
            sr_no: Value from Sr. No. column
            parent_code: Value from parent bag column
            
        Returns:
            True if this is a parent bag row
        """
        try:
            if sr_no is not None:
                float(sr_no)  # Can be 1.0 or 1
                
                # Check if parent code exists
                if parent_code and isinstance(parent_code, str) and parent_code.strip():
                    return True
        except (ValueError, TypeError):
            pass
        
        return False
    
    @staticmethod
    def is_bill_row(sr_no, bill_code) -> Tuple[bool, Optional[str]]:
        """
        Determine if a row represents a bill and extract bill number.
        
        Bill row criteria:
        - Sr. No. contains 'Bill' text (case-insensitive)
        - Bill code column contains the bill number
        
        Args:
            sr_no: Value from Sr. No. column
            bill_code: Value from bill code column
            
        Returns:
            Tuple of (is_bill_row, bill_number)
        """
        if sr_no and isinstance(sr_no, str) and 'bill' in sr_no.lower():
            # Extract bill number
            bill_number = str(bill_code).strip() if bill_code else None
            return True, bill_number
        
        return False, None
    
    @staticmethod
    def parse_excel_batch(file_storage: FileStorage) -> Tuple[List[Dict], List[str], Dict]:
        """
        Parse Excel file with batch pattern (parent bags followed by bill).
        
        Args:
            file_storage: Uploaded Excel file
            
        Returns:
            Tuple of (batches_list, error_list, stats_dict)
            
            batches_list format:
            [
                {
                    'bill_number': 'BILL001',
                    'parent_bags': ['SB12260', 'SB12248', ...],
                    'row_range': '2-10'
                },
                ...
            ]
        """
        if not EXCEL_AVAILABLE:
            return [], ["Excel support not available - openpyxl not installed"], {}
        
        try:
            file_storage.stream.seek(0)
            wb = load_workbook(file_storage.stream)
            ws = wb.active
            
            batches = []
            errors = []
            current_parents = []
            batch_start_row = None
            
            total_parents = 0
            total_bills = 0
            skipped_rows = 0
            
            # Process rows (skip header row 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip completely blank rows
                if not any(row):
                    skipped_rows += 1
                    continue
                
                sr_no = row[0] if len(row) > 0 else None
                parent_or_bill = row[1] if len(row) > 1 else None
                
                # Check if this is a parent bag row
                if ParentBillBatchImporter.is_parent_bag_row(sr_no, parent_or_bill):
                    if batch_start_row is None:
                        batch_start_row = row_num
                    
                    parent_code = str(parent_or_bill).strip()
                    current_parents.append(parent_code)
                    total_parents += 1
                
                # Check if this is a bill row
                is_bill, bill_number = ParentBillBatchImporter.is_bill_row(sr_no, parent_or_bill)
                if is_bill:
                    if not bill_number:
                        errors.append(f"Row {row_num}: Bill row found but bill number is blank/missing")
                        # Clear parents and start fresh
                        current_parents = []
                        batch_start_row = None
                        continue
                    
                    if not current_parents:
                        errors.append(f"Row {row_num}: Bill number '{bill_number}' found but no parent bags above it")
                        continue
                    
                    # Create batch
                    batches.append({
                        'bill_number': bill_number,
                        'parent_bags': current_parents.copy(),
                        'row_range': f"{batch_start_row}-{row_num}"
                    })
                    total_bills += 1
                    
                    # Reset for next batch
                    current_parents = []
                    batch_start_row = None
            
            # Handle orphaned parents at end of sheet
            if current_parents:
                errors.append(f"Warning: {len(current_parents)} parent bags found at end of sheet without a bill number")
                batches.append({
                    'bill_number': None,
                    'parent_bags': current_parents,
                    'row_range': f"{batch_start_row}-END",
                    'orphaned': True
                })
            
            stats = {
                'total_batches': len(batches),
                'total_parents': total_parents,
                'total_bills': total_bills,
                'skipped_rows': skipped_rows
            }
            
            return batches, errors, stats
        
        except Exception as e:
            logger.error(f"Excel batch parsing error: {str(e)}")
            return [], [f"Error parsing Excel file: {str(e)}"], {}
    
    @staticmethod
    def import_batches(db, batches: List[Dict], user_id: int) -> Tuple[int, int, int, List[str]]:
        """
        Import batches of parent bag → bill relationships.
        
        Args:
            db: Database session
            batches: List of batch dictionaries from parse_excel_batch
            user_id: ID of user performing import
            
        Returns:
            Tuple of (bills_created, links_created, parents_not_found, error_list)
        """
        from models import Bag, Bill, BillBag, BagType
        
        bills_created = 0
        links_created = 0
        parents_not_found = 0
        errors = []
        
        try:
            for batch_num, batch in enumerate(batches, 1):
                bill_number = batch['bill_number']
                parent_codes = batch['parent_bags']
                row_range = batch['row_range']
                is_orphaned = batch.get('orphaned', False)
                
                # Skip orphaned batches
                if is_orphaned:
                    errors.append(f"Batch {batch_num} (rows {row_range}): Skipped {len(parent_codes)} orphaned parents without bill")
                    continue
                
                # Start a savepoint for this batch
                savepoint = db.session.begin_nested()
                
                try:
                    # Check if bill already exists
                    bill = Bill.query.filter_by(bill_id=bill_number.upper()).first()
                    
                    if not bill:
                        # Create bill
                        bill = Bill(
                            bill_id=bill_number,
                            parent_bag_count=len(parent_codes),
                            created_by_id=user_id
                        )
                        db.session.add(bill)
                        db.session.flush()  # Get bill ID
                        bills_created += 1
                    else:
                        # Bill exists - we'll add parent bags to it
                        logger.info(f"Bill {bill_number} already exists, adding parent bags to it")
                    
                    # Link parent bags to bill
                    batch_links_created = 0
                    batch_parents_not_found = 0
                    
                    for parent_code in parent_codes:
                        # Find parent bag
                        parent_bag = Bag.query.filter_by(
                            qr_id=parent_code.upper(),
                            type=BagType.PARENT.value
                        ).first()
                        
                        if not parent_bag:
                            batch_parents_not_found += 1
                            errors.append(f"Batch {batch_num}: Parent bag '{parent_code}' not found in database")
                            continue
                        
                        # Check if link already exists
                        existing_link = BillBag.query.filter_by(
                            bill_id=bill.id,
                            bag_id=parent_bag.id
                        ).first()
                        
                        if not existing_link:
                            # Create link
                            bill_bag = BillBag(
                                bill_id=bill.id,
                                bag_id=parent_bag.id
                            )
                            db.session.add(bill_bag)
                            batch_links_created += 1
                    
                    # Update bill's parent bag count
                    actual_parent_count = BillBag.query.filter_by(bill_id=bill.id).count()
                    bill.parent_bag_count = actual_parent_count
                    
                    # Commit this batch
                    savepoint.commit()
                    db.session.flush()
                    
                    links_created += batch_links_created
                    parents_not_found += batch_parents_not_found
                    
                    logger.info(f"Batch {batch_num} ({row_range}): Linked bill '{bill_number}' to {batch_links_created} parent bags ({batch_parents_not_found} not found)")
                
                except Exception as batch_error:
                    savepoint.rollback()
                    error_msg = f"Batch {batch_num} (rows {row_range}, bill '{bill_number}'): {str(batch_error)}"
                    errors.append(error_msg)
                    logger.error(f"Batch import error: {error_msg}")
            
            # Final commit
            db.session.commit()
            
            # Log audit
            from audit_utils import log_audit
            log_audit(
                action='BATCH_IMPORT_PARENT_BILL',
                entity_type='bill',
                entity_id=None,
                details=f"Imported {len(batches)} batches: {bills_created} bills, {links_created} parent-bill links ({parents_not_found} parents not found)"
            )
            
            return bills_created, links_created, parents_not_found, errors
        
        except Exception as e:
            db.session.rollback()
            logger.error(f"Batch import error: {str(e)}")
            errors.append(f"Database error: {str(e)}")
            return bills_created, links_created, parents_not_found, errors


class MultiFileBatchProcessor:
    """
    Handles processing multiple Excel files in a single upload session.
    Generates detailed result reports with per-row status for large-scale imports.
    
    Features:
    - Memory-efficient streaming for large files
    - Per-row success/failure tracking
    - Downloadable Excel result files
    - Multi-file processing with combined reports
    """
    
    @staticmethod
    def generate_detailed_result_file(
        file_results: List[Dict],
        row_results: List[RowResult] = None,
        include_successful: bool = True
    ) -> bytes:
        """
        Generate comprehensive Excel result file with per-row status.
        
        Args:
            file_results: List of file-level processing results
            row_results: Optional list of individual row results
            include_successful: Whether to include successful rows (default True)
            
        Returns:
            Excel file as bytes
        """
        if not EXCEL_AVAILABLE:
            return b""
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Summary headers
        summary_headers = ['File Name', 'Status', 'Total Rows', 'Success', 'Errors', 
                          'Children Created', 'Links Created', 'Timestamp']
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Summary data
        row_num = 2
        for result in file_results:
            stats = result.get('stats', {})
            ws_summary.cell(row_num, 1, result.get('filename', 'Unknown'))
            ws_summary.cell(row_num, 2, result.get('status', 'Unknown'))
            ws_summary.cell(row_num, 3, stats.get('total_rows', 0))
            ws_summary.cell(row_num, 4, stats.get('children_created', 0) + stats.get('links_created', 0))
            ws_summary.cell(row_num, 5, stats.get('errors', 0))
            ws_summary.cell(row_num, 6, stats.get('children_created', 0))
            ws_summary.cell(row_num, 7, stats.get('links_created', 0))
            ws_summary.cell(row_num, 8, result.get('timestamp', ''))
            row_num += 1
        
        # Sheet 2: Detailed Results (per row)
        if row_results:
            ws_details = wb.create_sheet("Row Details")
            
            detail_headers = ['Row #', 'QR Code', 'Status', 'Message']
            for col, header in enumerate(detail_headers, 1):
                cell = ws_details.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Status colors
            status_colors = {
                RowResult.SUCCESS: "C6EFCE",  # Light green
                RowResult.CHILD_CREATED: "C6EFCE",  # Light green
                RowResult.LINKED: "C6EFCE",  # Light green
                RowResult.DUPLICATE: "FFEB9C",  # Light yellow
                RowResult.SKIPPED: "DDDDDD",  # Light gray
                RowResult.ERROR: "FFC7CE",  # Light red
            }
            
            row_num = 2
            for result in row_results:
                # Skip successful rows if not requested
                if not include_successful and result.status in [RowResult.SUCCESS, RowResult.CHILD_CREATED, RowResult.LINKED]:
                    continue
                
                ws_details.cell(row_num, 1, result.row_num)
                ws_details.cell(row_num, 2, result.qr_code[:50] if result.qr_code else '')
                ws_details.cell(row_num, 3, result.status)
                ws_details.cell(row_num, 4, result.message[:200] if result.message else '')
                
                # Apply status color
                fill_color = status_colors.get(result.status, "FFFFFF")
                for col in range(1, 5):
                    ws_details.cell(row_num, col).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
                
                row_num += 1
                
                # Limit to MAX_ERRORS_PER_FILE to prevent huge files
                if row_num > MAX_ERRORS_PER_FILE + 1:
                    ws_details.cell(row_num, 1, "...")
                    ws_details.cell(row_num, 4, f"Truncated - {len(row_results) - MAX_ERRORS_PER_FILE} more rows not shown")
                    break
        
        # Sheet 3: Successes Only
        if row_results and include_successful:
            ws_success = wb.create_sheet("Successes")
            success_headers = ['Sheet', 'Row #', 'Parent Bag', 'Child Bag', 'Action']
            for col, header in enumerate(success_headers, 1):
                cell = ws_success.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            row_num = 2
            success_count = 0
            for result in row_results:
                if result.status in [RowResult.SUCCESS, RowResult.CHILD_CREATED, RowResult.LINKED, RowResult.PARENT_CREATED]:
                    sheet_name = result.details.get('sheet', '') if result.details else ''
                    parent_qr = result.details.get('parent_qr', '') if result.details else ''
                    child_qr = result.details.get('child_qr', result.qr_code) if result.details else result.qr_code
                    
                    ws_success.cell(row_num, 1, sheet_name)
                    ws_success.cell(row_num, 2, result.row_num)
                    ws_success.cell(row_num, 3, parent_qr)
                    ws_success.cell(row_num, 4, child_qr[:50] if child_qr else '')
                    ws_success.cell(row_num, 5, result.status)
                    
                    # Apply green fill
                    for col in range(1, 6):
                        ws_success.cell(row_num, col).fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    
                    row_num += 1
                    success_count += 1
                    
                    # Cap at 5000 successes to save memory
                    if success_count >= 5000:
                        ws_success.cell(row_num, 1, "...")
                        ws_success.cell(row_num, 5, f"Capped at 5000 - more successes not shown")
                        break
        
        # Sheet 4: Errors Only
        ws_errors = wb.create_sheet("Errors")
        error_headers = ['Sheet', 'Row #', 'QR Code', 'Error Message']
        for col, header in enumerate(error_headers, 1):
            cell = ws_errors.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row_num = 2
        # Add file-level errors
        for result in file_results:
            for error in result.get('errors', [])[:100]:  # Limit per file
                ws_errors.cell(row_num, 1, '')
                ws_errors.cell(row_num, 2, '')
                ws_errors.cell(row_num, 3, '')
                ws_errors.cell(row_num, 4, str(error)[:500])
                # Apply red fill
                for col in range(1, 5):
                    ws_errors.cell(row_num, col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                row_num += 1
        
        # Add row-level errors
        if row_results:
            for result in row_results:
                if result.status == RowResult.ERROR:
                    sheet_name = result.details.get('sheet', '') if result.details else ''
                    ws_errors.cell(row_num, 1, sheet_name)
                    ws_errors.cell(row_num, 2, result.row_num)
                    ws_errors.cell(row_num, 3, result.qr_code[:50] if result.qr_code else '')
                    ws_errors.cell(row_num, 4, result.message[:500] if result.message else '')
                    # Apply red fill
                    for col in range(1, 5):
                        ws_errors.cell(row_num, col).fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )
                    row_num += 1
                    
                    if row_num > MAX_ERRORS_PER_FILE + 1:
                        break
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    @staticmethod
    def generate_error_report(results: List[Dict]) -> bytes:
        """
        Generate an Excel error report from processing results.
        (Legacy method - kept for backward compatibility)
        
        Args:
            results: List of file processing results
            
        Returns:
            Excel file as bytes
        """
        if not EXCEL_AVAILABLE:
            return b""
        
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Errors"
        
        # Add headers
        headers = ['File Name', 'Import Type', 'Status', 'Row/Batch', 'Error Details', 'Timestamp']
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)
            ws.cell(1, col).font = ws.cell(1, col).font.copy(bold=True)
        
        # Add error data
        row = 2
        for result in results:
            filename = result.get('filename', 'Unknown')
            import_type = result.get('import_type', 'Unknown')
            status = result.get('status', 'Unknown')
            timestamp = result.get('timestamp', '')
            
            # Add main file status
            ws.cell(row, 1, filename)
            ws.cell(row, 2, import_type)
            ws.cell(row, 3, status)
            ws.cell(row, 4, '')
            ws.cell(row, 5, result.get('summary', ''))
            ws.cell(row, 6, timestamp)
            row += 1
            
            # Add detailed errors if any
            errors = result.get('errors', [])
            for error in errors:
                ws.cell(row, 1, filename)
                ws.cell(row, 2, import_type)
                ws.cell(row, 3, 'Error')
                ws.cell(row, 4, '')
                ws.cell(row, 5, error)
                ws.cell(row, 6, timestamp)
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 7):
            max_length = 0
            column_letter = ws.cell(1, col).column_letter
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    @staticmethod
    def process_child_parent_files_streaming(
        files: List[FileStorage], 
        user_id: int, 
        dispatch_area: Optional[str] = None,
        auto_create_parents: bool = False
    ) -> Tuple[List[Dict], List[RowResult], bool]:
        """
        Process multiple child-parent batch import files using streaming for large files.
        Returns per-row results for detailed reporting.
        
        Args:
            files: List of uploaded Excel files
            user_id: ID of user performing import
            dispatch_area: Optional dispatch area filter
            auto_create_parents: If True, automatically create parent bags if missing
            
        Returns:
            Tuple of (file_results_list, all_row_results, has_errors)
        """
        from datetime import datetime
        
        file_results = []
        all_row_results = []
        has_errors = False
        
        for file in files:
            filename = file.filename or 'unknown.xlsx'
            result = {
                'filename': filename,
                'import_type': 'Child → Parent',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'errors': [],
                'summary': '',
                'stats': {}
            }
            
            try:
                # Reset stream position
                file.stream.seek(0)
                
                # Use streaming importer for memory efficiency
                logger.info(f"Processing {filename} with streaming importer (auto_create_parents={auto_create_parents})")
                stats, row_results = LargeScaleChildParentImporter.process_file_streaming(
                    file_storage=file,
                    user_id=user_id,
                    dispatch_area=dispatch_area,
                    auto_create_parents=auto_create_parents
                )
                
                result['stats'] = stats
                all_row_results.extend(row_results)
                
                # Check for errors
                if stats.get('errors', 0) > 0 or stats.get('parents_not_found', 0) > 0:
                    has_errors = True
                    result['status'] = 'Partial Success'
                else:
                    result['status'] = 'Success'
                
                result['summary'] = (
                    f"{stats.get('batches_processed', 0)} batches, "
                    f"{stats.get('children_created', 0)} children created, "
                    f"{stats.get('links_created', 0)} links created"
                )
                
                if stats.get('parents_not_found', 0) > 0:
                    result['summary'] += f", {stats.get('parents_not_found', 0)} parents not found"
                if stats.get('errors', 0) > 0:
                    result['summary'] += f", {stats.get('errors', 0)} errors"
                
                # Extract error messages from row results
                error_rows = [r for r in row_results if r.status == RowResult.ERROR]
                if error_rows:
                    result['errors'] = [f"Row {r.row_num}: {r.message}" for r in error_rows[:50]]
                
            except Exception as e:
                logger.error(f"Error processing file {filename}: {str(e)}")
                result['status'] = 'Failed'
                result['summary'] = f'Fatal error: {str(e)}'
                result['errors'].append(str(e))
                has_errors = True
            
            file_results.append(result)
        
        return file_results, all_row_results, has_errors
    
    @staticmethod
    def process_child_parent_files(files: List[FileStorage], user_id: int, dispatch_area: Optional[str] = None) -> Tuple[List[Dict], bool]:
        """
        Process multiple child-parent batch import files.
        (Legacy method - uses non-streaming for backward compatibility)
        
        Args:
            files: List of uploaded Excel files
            user_id: ID of user performing import
            dispatch_area: Optional dispatch area filter
            
        Returns:
            Tuple of (results_list, has_errors)
        """
        from datetime import datetime
        
        results = []
        has_errors = False
        
        for file in files:
            filename = file.filename or 'unknown.xlsx'
            result = {
                'filename': filename,
                'import_type': 'Child → Parent',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'errors': [],
                'summary': ''
            }
            
            try:
                # Reset stream position
                file.stream.seek(0)
                
                # Parse batches
                batches, parse_errors, stats = ChildParentBatchImporter.parse_excel_batch(file)
                
                # Add parse errors
                if parse_errors:
                    result['errors'].extend(parse_errors[:20])  # Limit to first 20 errors
                    has_errors = True
                
                if not batches:
                    result['status'] = 'Failed'
                    result['summary'] = 'No valid batches found'
                    has_errors = True
                    results.append(result)
                    continue
                
                # Import batches
                parents_created, children_created, links_created, parents_not_found, import_errors = ChildParentBatchImporter.import_batches(
                    db, batches, user_id, dispatch_area
                )
                
                # Add import errors
                if import_errors:
                    result['errors'].extend(import_errors[:20])  # Limit to first 20 errors
                    has_errors = True
                
                # Set status
                if import_errors or parse_errors:
                    result['status'] = 'Partial Success'
                else:
                    result['status'] = 'Success'
                
                result['summary'] = f'{len(batches)} batches, {parents_created} parents, {children_created} children, {links_created} links'
                if parents_not_found > 0:
                    result['summary'] += f', {parents_not_found} parents not found'
                
            except Exception as e:
                logger.error(f"Error processing file {filename}: {str(e)}")
                result['status'] = 'Failed'
                result['summary'] = f'Fatal error: {str(e)}'
                result['errors'].append(str(e))
                has_errors = True
            
            results.append(result)
        
        return results, has_errors
    
    @staticmethod
    def process_parent_bill_files(files: List[FileStorage], user_id: int) -> Tuple[List[Dict], bool]:
        """
        Process multiple parent-bill batch import files.
        
        Args:
            files: List of uploaded Excel files
            user_id: ID of user performing import
            
        Returns:
            Tuple of (results_list, has_errors)
        """
        from datetime import datetime
        
        results = []
        has_errors = False
        
        for file in files:
            filename = file.filename or 'unknown.xlsx'
            result = {
                'filename': filename,
                'import_type': 'Parent → Bill',
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'errors': [],
                'summary': ''
            }
            
            try:
                # Reset stream position
                file.stream.seek(0)
                
                # Parse batches
                batches, parse_errors, stats = ParentBillBatchImporter.parse_excel_batch(file)
                
                # Add parse errors
                if parse_errors:
                    result['errors'].extend(parse_errors[:20])  # Limit to first 20 errors
                    has_errors = True
                
                if not batches:
                    result['status'] = 'Failed'
                    result['summary'] = 'No valid batches found'
                    has_errors = True
                    results.append(result)
                    continue
                
                # Import batches
                bills_created, links_created, parents_not_found, import_errors = ParentBillBatchImporter.import_batches(
                    db, batches, user_id
                )
                
                # Add import errors
                if import_errors:
                    result['errors'].extend(import_errors[:20])  # Limit to first 20 errors
                    has_errors = True
                
                # Set status
                if import_errors or parse_errors:
                    result['status'] = 'Partial Success'
                else:
                    result['status'] = 'Success'
                
                result['summary'] = f'{len(batches)} batches, {bills_created} bills, {links_created} links'
                if parents_not_found > 0:
                    result['summary'] += f', {parents_not_found} parents not found'
                
            except Exception as e:
                logger.error(f"Error processing file {filename}: {str(e)}")
                result['status'] = 'Failed'
                result['summary'] = f'Fatal error: {str(e)}'
                result['errors'].append(str(e))
                has_errors = True
            
            results.append(result)
        
        return results, has_errors
